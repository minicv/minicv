import math
import os
import threading

import openpyxl
import requests

PROJ_ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILE_DIR = os.path.join(PROJ_ROOT_DIR, 'files')
IMG_DIR = os.path.join(FILE_DIR, 'images')
TOUTIAO_ANNO_FILE_PATH = os.path.join(FILE_DIR, 'toutiao_content_image.xlsx')
OUTPUT_XLSX_FILE_PATH = os.path.join(FILE_DIR, 'article_attr.xlsx')
TOUTIAO_ANNO_TITLES = ('article_title', 'article_class', 'tags', 'article_url',
                       'img_urls', 'article_content')
OUTPUT_XLSX_TITLES = ('article_id', 'img_urls', 'img_paths', 'has_content',
                      'contain_car', 'contain_human', 'human_tag',
                      'object tags')
MAX_IMG_PER_ARTICLE = 5
NUM_THREADS = 40


def down_img(url, save_path):
    try:
        rsp = requests.get(url)
        if rsp.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(rsp.content)
        return True

    except Exception as e:
        print(e)
        return False


def process_article(article_id, input_xlsx_sheet, output_xlsx_sheet):
    """Performs image downloading and data processing for an article.

    :param article_id: Index for the article in the anno file (int).
    :param input_xlsx_sheet: The sheet to read (openpyxl worksheet).
    :param output_xlsx_sheet: The sheet to write (openpyxl worksheet).
    """
    # row and column index start from 1
    row_index = article_id + 1
    article_class = input_xlsx_sheet.cell(
        row=row_index, column=2).value.strip()
    img_urls = input_xlsx_sheet.cell(row=row_index, column=5).value
    article_content = input_xlsx_sheet.cell(row=row_index, column=6).value

    if not img_urls:
        img_url_list = []
    else:
        num_imgs = min(len(img_urls), MAX_IMG_PER_ARTICLE)
        img_url_list = img_urls.strip().split(',')[:num_imgs]
    img_path_list = []
    has_content = (article_content is not None)
    contain_car = '汽车' in article_class
    article_dir = os.path.join(IMG_DIR, 'article_' + str(article_id))
    non_prefix_article_dir = os.path.join(
        os.path.basename(IMG_DIR), 'article_' + str(article_id))

    if not os.path.exists(article_dir):
        os.makedirs(article_dir)

    for i, img_url in enumerate(img_url_list):
        img_save_path = os.path.join(article_dir, 'img_' + str(i) + '.jpg')
        non_prefix_save_path = os.path.join(non_prefix_article_dir,
                                            'img_' + str(i) + '.jpg')

        if os.path.exists(img_save_path):
            existp = True
        else:
            existp = down_img(img_url, img_save_path)

        img_path_list.append(non_prefix_save_path if existp else '')

    img_paths = ','.join(img_path_list)

    # row_index + 1 because the first row of the sheet is the title row
    output_xlsx_sheet.cell(row=row_index + 1, column=1).value = article_id
    output_xlsx_sheet.cell(row=row_index + 1, column=2).value = img_urls
    output_xlsx_sheet.cell(row=row_index + 1, column=3).value = img_paths
    output_xlsx_sheet.cell(
        row=row_index + 1, column=4).value = int(has_content)
    output_xlsx_sheet.cell(
        row=row_index + 1, column=5).value = int(contain_car)
    output_xlsx_sheet.cell(row=row_index + 1, column=6).value = ''
    output_xlsx_sheet.cell(row=row_index + 1, column=7).value = ''
    output_xlsx_sheet.cell(row=row_index + 1, column=8).value = ''


def process_split(split, input_xlsx_sheet, output_xlsx_sheet):
    """Processes a list of articles in the split.

    :param split: The indices of the articles (list).
    :param input_xlsx_sheet: The xlsx sheet to read (openpyxl worksheet).
    :param output_xlsx_sheet: The xlsx sheet to write (openpyxl worksheet).
    """
    for article_id in split:
        process_article(article_id, input_xlsx_sheet, output_xlsx_sheet)


def main():
    # prepare the input and output excel files
    toutiao_anno_workbook = openpyxl.load_workbook(TOUTIAO_ANNO_FILE_PATH)
    toutiao_anno_sheet = toutiao_anno_workbook.active
    output_xlsx_workbook = openpyxl.Workbook()
    output_xlsx_sheet = output_xlsx_workbook.active

    for i, title in enumerate(OUTPUT_XLSX_TITLES):
        # row and column index starts from 1
        output_xlsx_sheet.cell(row=1, column=i + 1).value = title

    # TODO: remove '[:100]' to access the full list
    article_ids = list(range(0, toutiao_anno_sheet.max_row))[:80000]
    n_threads = NUM_THREADS
    split_size = math.ceil(len(article_ids) / n_threads)
    splits = []

    for i_thread in range(n_threads):
        start = i_thread * split_size
        end = (i_thread + 1) * split_size
        if end > len(article_ids):
            end = len(article_ids)
        split = article_ids[start:end]
        splits.append(split)

    threads = []
    for split in splits:
        thread = threading.Thread(
            target=process_split,
            args=(split, toutiao_anno_sheet, output_xlsx_sheet))
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()

    output_xlsx_workbook.save(OUTPUT_XLSX_FILE_PATH)


if __name__ == '__main__':
    main()
